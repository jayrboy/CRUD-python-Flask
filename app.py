from flask import Flask, render_template, url_for, redirect, session, request, send_file, jsonify
import os
import mysql.connector as mysql
import hashlib
from openpyxl import Workbook, load_workbook
import pandas as pd

app = Flask(__name__)

# กำหนดรหัสลับ สำหรับป้องกันจากผู้นักพัฒนาที่จะ hack แก้ไข session cookie และ ข้อมูลอื่นๆ
app.secret_key = '5f352379324c22463451387a0aec5d2f'

conn = mysql.connect(
    host = 'localhost',
    user = 'root',
    passwd = '',
    port = 3306,
    database = 'first_project'
)

#๊ Upload Files into PATH : /static/excel
app.config["UPLOAD_FOLDER1"]="static/excel"

# Define images limit = 4MB
app.config['MAX_CONTENT_LENGTH'] = 4 * 1000 * 1000 

#๊ Upload Files into PATH : /static/picture
UPLOAD_PIC_FOLDER = os.path.join(os.path.dirname(__file__), "static/picture")

# Import Folder from HTML-CSS-JS
template_folder = os.path.join(os.path.dirname(__file__), "templates/")
app.static_folder = 'static'
app.static_url_path = '/static'

@app.route('/', methods=["GET","POST"])
def index():
    session['user'] = ''
    session['audit'] = False
    return render_template("sign-in.html")

@app.route('/sign-up', methods=["GET","POST"])
def sign_up():
    return render_template("sign-up.html")

@app.route('/validate-sign-up', methods=["GET","POST"])
def validate_sign_up():
    # ส่งข้อมูล input ทุกอย่าง และส่ง output ไปที่ /validate-sign-up
    name = request.form['fname']
    user = request.form['user']
    password = request.form['password']
    cfpassword = request.form['cfpassword']
    mobile = request.form['mobile']
    birthday = request.form['birthday']
    gender = request.form['gender']


    if name != "" and user != "" and password !="" and password == cfpassword and mobile != ""and birthday != ""and gender != "":
        # encypt: Unknow word
        encypt_passwd = hashlib.md5(password.encode()).hexdigest()
        conn.reconnect()
        cur = conn.cursor() #Connected MySQL Server
        insert_sql = """
            INSERT INTO username(username,password,fullname,audit,mobile,birthday,gender)
            VALUES(%s,%s,%s,%s,%s,%s,%s)
        """
        val = (user,encypt_passwd,name,1,mobile,birthday,gender)
        cur.execute(insert_sql, val)
        conn.commit()
        conn.close()
        return render_template('sign-up-success.html')
        
        # ถ้ากรอกไม่ครบจะกลับไปที่หน้า /sign-up
    else:
        return render_template('sign-up.html')

@app.route('/validate-sign-in', methods=["GET","POST"])
def validate_sign_in():
    user = request.form['user']
    passwd = request.form['password']

    if user !='' and passwd !='':
        sql = '''
            SELECT password FROM username 
            WHERE username=%s
        '''
        val = (user,)
        # reconnect ก่อนการเข้าถึง database ก่อน ป้องกัน Server ค้าง
        conn.reconnect()
        cur = conn.cursor()
        cur.execute(sql, val)
        data = cur.fetchone()
        conn.close()
        encypt_passwd = hashlib.md5(passwd.encode()).hexdigest()
        # data[0] = ('password',) เปรียบเทียบ กับ password ใน database
        if data[0] == encypt_passwd:
            session['user'] = user
            session['audit'] = True
            return render_template('sign-in-success.html')
        else:
            return render_template('sign-in.html')
    else:
        return render_template('sign-in.html')

@app.route('/main-program', methods=["GET","POST"])
def main_program():
    if session['audit'] == True:
        return render_template('main.html')
    else:
        return redirect('/')

@app.route('/sign-out', methods=["GET","POST"])
def sign_out():
    session.pop('user')
    session.pop('audit')
    return redirect('/')

@app.route('/user', methods=["GET","POST"])
def user():
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            SELECT username,fullname,mobile,birthday,picture
            FROM username
        '''
    cur.execute(sql)
    data = cur.fetchall()
    conn.close()
    return render_template('user.html', users=data)

@app.route('/user-add', methods=["GET","POST"])
def user_add():
    return render_template('user-add.html')

@app.route('/user-add-post', methods=["GET","POST"])
def user_add_post():
    name = request.form['fname']
    user = request.form['user']
    password = request.form['password']
    cfpassword = request.form['cfpassword']
    mobile = request.form['mobile']
    birthday = request.form['birthday']
    gender = request.form['gender']
    picture = request.files['picture']

    encrypt_pass = hashlib.md5(password.encode()).hexdigest()

    picture.save(os.path.join(UPLOAD_PIC_FOLDER, picture.filename))

    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            INSERT INTO username(username,password,fullname,audit,mobile,birthday,gender,picture)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
        '''
    val = (user,encrypt_pass,name,1,mobile,birthday,gender,picture.filename)
    cur.execute(sql, val)
    conn.commit()
    conn.close()
    return redirect('/user')

@app.route('/user-delete/<user>', methods=["GET","POST"])
def user_delete(user):
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            DELETE FROM username WHERE username=%s
        '''
    val = (user,)
    cur.execute(sql, val)
    conn.commit()
    conn.close()
    return redirect('/user')

@app.route('/user-edit/<user>', methods=["GET","POST"])
def user_edit(user):
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            SELECT username,fullname,mobile,birthday,gender,picture FROM username WHERE username=%s
        '''
    val = (user,)
    cur.execute(sql, val)
    data = cur.fetchone()
    conn.close()
    return render_template('user-edit.html', users=data)

@app.route('/user-edit-post', methods=["GET","POST"])
def user_edit_post():
    name = request.form['fname']
    user = request.form['user']    
    password = request.form['password']
    cfpassword = request.form['cfpassword']
    mobile = request.form['mobile']
    birthday = request.form['birthday']
    gender = request.form['gender'] 
    picture = request.files['picture']

    
    if password:
        encrypt_pass = hashlib.md5(password.encode()).hexdigest()

    if picture:
        picture.save(os.path.join(UPLOAD_PIC_FOLDER, picture.filename))

    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            UPDATE username SET fullname=%s, password=%s, mobile=%s, birthday=%s, gender=%s, picture=%s
            WHERE username=%s
        '''
    val = (name,encrypt_pass,mobile,birthday,gender,picture.filename,user)
    cur.execute(sql, val)
    conn.commit()
    conn.close()
    return redirect('/user')

@app.route('/user-download', methods=["GET"])
def user_download():
    file = 'username.xlsx'
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            SELECT * 
            FROM username
        '''
    cur.execute(sql)
    username = cur.fetchall()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['id', 'username', 'password', 'fullname', 'audit', 'mobile', 'birthday', 'gender', 'picture'])

    for u in username:
        print(u)
        sheet.append(u)

    workbook.save(file)
    workbook.close()

    conn.close()
    return send_file(file, as_attachment=True)

@app.route('/product', methods=["GET","POST"])
def product():
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            SELECT *
            FROM product
        '''
    cur.execute(sql)
    data = cur.fetchall()
    conn.close()
    return render_template('product.html', product=data)

@app.route('/product-add', methods=["GET","POST"])
def product_add():
    return render_template('product-add.html')

@app.route('/product-add-post', methods=["POST"])
def product_add_post():
    product = request.form['product']
    types = request.form['types']
    price = request.form['price']
    details = request.form['details']
    picture = request.files['picture']

    if picture:
        picture.save(os.path.join(UPLOAD_PIC_FOLDER, picture.filename))

    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            INSERT INTO product(name,types,price,details,picture)
            VALUES(%s,%s,%s,%s,%s)
        '''
    val = (product,types,price,details,picture.filename)
    cur.execute(sql, val)
    conn.commit()
    conn.close()
    return redirect('/product')

@app.route('/product-delete/<id>', methods=["GET"])
def product_delete(id):
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            DELETE 
            FROM product 
            WHERE id=%s
        '''
    val = (id,)
    cur.execute(sql, val)
    conn.commit()
    conn.close()
    return redirect('/product')

@app.route('/product-edit/<product>', methods=["GET","POST"])
def product_edit(product):
    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            SELECT id,name,types,price,details,picture,picture FROM product WHERE id=%s
        '''
    val = (product,)
    cur.execute(sql, val)
    data = cur.fetchone()
    conn.close()
    return render_template('product-edit.html', product=data)

@app.route('/product-edit-post', methods=["POST"])
def product_edit_post():
    id = request.form['id']
    product = request.form['product']
    types = request.form['types']
    price = request.form['price']
    details = request.form['details']
    picture = request.files['picture']

    conn.reconnect()
    cur = conn.cursor()
    sql = '''
            UPDATE product SET name=%s,types=%s, price=%s, details=%s, picture=%s
            WHERE id=%s
        '''
    val = (product,types,price,details,picture.filename,id)
    cur.execute(sql, val)
    conn.commit()
    conn.close()
    return redirect('/product')

@app.route('/product-download', methods=["GET"])
def product_download():
    path = 'assets.xlsx'
    # Connect Database
    conn.reconnect()
    # - ส่งคำสั่ง SQL ไปให้ MySQL ทำการโหลดข้อมูล
    # - Python จะรับข้อมูลทั้งหมดมาเป็น List ผ่านคำสั่ง fetchall()
    cur = conn.cursor()
    sql = '''
            SELECT * 
            FROM product
        '''
    cur.execute(sql)
    products = cur.fetchall()
    # Excel
    # - สร้างไฟล์ใหม่ สร้างชีท และใส่แถวสำหรับเป็นหัวข้อตาราง
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['id', 'name', 'types', 'price', 'details', 'picture'])
    # - ใส่ข้อมูลทีละอัน เพิ่มลงไปทีละแถว
    for p in products:
        print(p)
        sheet.append(p)
    # - Export ไฟล์ Excel
    workbook.save('assets.xlsx')
    workbook.close()
    conn.close()
    return send_file(path, as_attachment=True)

@app.route('/product-upload', methods=["GET","POST"])
def product_upload():
    if request.method == 'POST':
        upload_file = request.files['upload_excel']
        workbook = load_workbook(upload_file)
        sheet = workbook.active
        
        if upload_file.filename != '':
            file_path = os.path.join(app.config["UPLOAD_FOLDER1"], upload_file.filename)
            upload_file.save(file_path)
            data=pd.read_excel(upload_file)
            
            values = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(row)
                values.append(row)
            
            conn.reconnect()
            cur = conn.cursor()
            sql = '''
                INSERT INTO product (name,types,price,details,picture)
                VALUES (%s, %s, %s, %s, %s)
                '''
            cur.executemany(sql, values)
            conn.commit()
            conn.close()
        
        return render_template("product-upload.html",data=data.to_html(index=False).replace('<th>','<th class="container" style="text-align:center;background-color:#000;color:#fff">'))
    return render_template("product-upload.html")

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
















